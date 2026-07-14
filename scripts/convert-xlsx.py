"""
Converts the AIESEC in Indore opportunity spreadsheet into data/opportunities.json.
Usage: python3 scripts/convert-xlsx.py path/to/sheet.xlsx
Expected columns: ID, Title, Organization, Status, Link
Program type is derived from the EXPA link (/opportunity/<type>/<id>).
"""
import sys, json, re
from openpyxl import load_workbook

TYPE_MAP = {
    "global-volunteer": "Volunteer",
    "global-talent": "Talent",
    "global-teacher": "Teacher",
}

def slugify(text):
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")

def convert(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    out = []
    for r in rows[1:]:
        if not r or r[idx.get("ID", 0)] is None:
            continue
        raw_id = r[idx["ID"]]
        opp_id = int(raw_id) if isinstance(raw_id, float) else raw_id
        title = (r[idx["Title"]] or "").strip()
        org = (r[idx["Organization"]] or "").strip()
        status = (r[idx["Status"]] or "").strip()
        link = (r[idx["Link"]] or "").strip()

        m = re.search(r"/opportunity/([a-z0-9-]+)/", link)
        type_slug = m.group(1) if m else "unknown"
        program_type = TYPE_MAP.get(type_slug, "Other")

        out.append({
            "id": opp_id,
            "slug": f"{slugify(title)}-{opp_id}",
            "title": title,
            "organization": org,
            "programType": program_type,
            "status": status,
            "expaLink": link,
        })

    return out

if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/IGTae_Search_Tool.xlsx"
    data = convert(src)
    out_path = "src/data/opportunities.json"
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
    print(f"Wrote {len(data)} opportunities to {out_path}")
